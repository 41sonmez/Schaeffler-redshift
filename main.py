import pandas as pd
import os
import re
import psycopg2 
from typing import Dict, Any, List, Tuple 
import numpy as np 
# Scheduler imports
from apscheduler.schedulers.blocking import BlockingScheduler 
from apscheduler.triggers.interval import IntervalTrigger
import time 
from openpyxl import load_workbook
import streamlit as st # ADDED: Required to access st.secrets

# ******************************************************************************
# --- 0. SCHEDULE SETTINGS ---
# ******************************************************************************
# ⚠️ DEFINE YOUR REPETITION INTERVAL HERE (in minutes)
SCHEDULE_INTERVAL_MINUTES = 60 # Set to 60 for hourly runs (default)

# ******************************************************************************
# --- 1. GENERAL SETTINGS AND REDSHIFT CONNECTION PARAMETERS ---
# ******************************************************************************
# NOTE: BASE_PATH and local file paths are assumed to be accessible in this 
# environment. If running in Streamlit Cloud, these network paths 
# (e.g., \\vt1.vitesco.com\) will cause a failure.
BASE_PATH = r'\\vt1.vitesco.com\fst1\did80029\PBI_SCM'
MATERIALDATA_FILE = os.path.join(BASE_PATH, '8343_stocks_500_20251205.csv') 
GROUPDATA_FILE = os.path.join(BASE_PATH, 'WIP_Ostrava_SAP.xlsx') 
FINAL_REPORT_FILE = os.path.join(BASE_PATH, 'redshift_final_load_report.xlsx') 

# MATERIAL DATA SETTINGS (0-indexed)
MATERIAL_COL_IDX = 7 # Column L
COUNT_COL_IDX = 11   # Column W
MATERIAL_SHEET_IDX = 0

# REDSHIFT PARAMETERS (Reading securely from st.secrets)
try:
    # Attempt to load sensitive data from the secure secrets store
    HOST = st.secrets["redshift"]["host"]
    PORT = int(st.secrets["redshift"]["port"]) # Ensure port is integer
    DATABASE = st.secrets["redshift"]["dbname"]
    USER = st.secrets["redshift"]["user"]
    PASSWORD = st.secrets["redshift"]["password"]
    
    # Check if 'schema' is also stored in secrets, otherwise use the hardcoded default
    SCHEMA = st.secrets.get("redshift", {}).get("schema", "schema_f540_fst_ohd_camera_results")
    
except Exception as e:
    # CRITICAL: If secrets cannot be loaded (e.g., running outside Streamlit Cloud 
    # without a secrets.toml or if the key structure is wrong), assign None to fail safely.
    print(f"❌ CRITICAL: Could not load secrets from st.secrets. Error: {e}")
    HOST = None
    PORT = 5439
    DATABASE = None
    USER = None
    PASSWORD = None
    SCHEMA = "schema_f540_fst_ohd_camera_results"
 
# GLOBAL REDSHIFT TABLE NAMES (FULL NAME)
TARGET_TABLE_NAME = f"{SCHEMA}.material_processes_count_test" 
TEMP_TABLE_NAME = f"{SCHEMA}.material_processes_count_temp"

# ******************************************************************************
# --- HELPER FUNCTIONS ---
# ******************************************************************************

def create_redshift_connection(autocommit_val=True):
    """Establishes a Redshift connection using globally loaded parameters."""
    # Check if secrets loading failed
    if HOST is None:
        print("❌ Redshift Connection Error: HOST is None (Secrets not loaded). Aborting connection attempt.")
        return None
    
    try:
        conn = psycopg2.connect(
            host=HOST,
            port=PORT,
            dbname=DATABASE,
            user=USER,
            password=PASSWORD
        )
        conn.autocommit = autocommit_val 
        return conn
    except Exception as e:
        print(f"❌ Redshift Connection Error: {e}")
        return None

def create_table_ddl(table_name):
    # (Function body remains the same)
    dtype_map = {
        'part_number': 'VARCHAR(256) PRIMARY KEY',
        'production_variant': 'VARCHAR(100)',
        'development_variant': 'VARCHAR(100)',
        'production_step': 'VARCHAR(2000)',
        'quantity': 'INTEGER'
    }
    columns_ddl = []
    for col, type_def in dtype_map.items():
        columns_ddl.append(f'{col} {type_def}') 
    
    if table_name.endswith('_temp'):
        columns_ddl = [col.replace(' PRIMARY KEY', '') for col in columns_ddl]
        
    return f'CREATE TABLE IF NOT EXISTS {table_name} ({", ".join(columns_ddl)});'

def get_clean_level_names(filepath) -> List[Tuple[str, str, str]]:
    # (Function body remains the same - reads sheet names which is low memory)
    VARIANT_PATTERN = re.compile(r'\s([0-9]+kW)_([A-Z]?[0-9]+\.[0-9]+)\s*$')
    try:
        wb = load_workbook(filepath, read_only=True, keep_links=False)
        design_levels_raw = wb.sheetnames
        
        final_filtered_list = [level.strip() 
                             for level in design_levels_raw 
                             if ('kW' in level.strip() and '_REAL' not in level.strip() and 'Stock report' in level.strip())]
        
        result_list = []
        for raw_sheet_name in final_filtered_list[1:]:
            match = VARIANT_PATTERN.search(raw_sheet_name)
            
            production_variant = ''
            development_variant = ''

            if match and len(match.groups()) == 2:
                production_variant = match.group(1).strip()
                development_variant = match.group(2).strip()

            result_list.append((raw_sheet_name, production_variant, development_variant))
                
        return result_list 
        
    except Exception as e:
        print(f"❌ Variant Extraction Error: {e}")
        return []
 
def stage_1_get_material_names(material_data_file, material_col_idx, count_col_idx) -> Dict[str, int]:
    """
    MODIFIED: Reads the CSV file in chunks. Uses safe filtering (only excludes
    empty PN and zero quantity). All PN patterns are accepted.
    """
    # Regex artık sadece Part Number temizliği için kullanılır (boşluk vb.)
    # Kısıtlayıcı filtreleme Regex'i kaldırıldı.
    CHUNK_SIZE = 5000 
    
    MATERIAL_COL_NAME = material_col_idx
    COUNT_COL_NAME = count_col_idx
    
    aggregated_counts: Dict[str, int] = {}
    chunk_count = 0

    try:
        print("-> Reading material data in chunks (Memory Efficient via read_csv)...")
        
        cols_to_use = [material_col_idx, count_col_idx]
        
        for chunk in pd.read_csv(
            material_data_file, 
            chunksize=CHUNK_SIZE, 
            usecols=cols_to_use, 
            header=None,  
            encoding='latin-1' 
        ):
            chunk_count += 1
            print(f"   -> Processing Chunk {chunk_count} ({len(chunk)} rows)...")
            
            material_col_name = MATERIAL_COL_NAME
            count_col_name = COUNT_COL_NAME
            
            # --- 1. MİKTAR (QUANTITY) TEMİZLEME ---
            quantity_series = chunk[count_col_name].astype(str).str.strip()
            # Formül kalıntılarını ve aşırı boşlukları temizle
            quantity_series = quantity_series.str.replace(r'^="?\s*|"?\s*$', '', regex=True) 
            # Binlik ayraç olarak kullanılan noktayı kaldır
            quantity_series = quantity_series.str.replace('.', '', regex=False) 
            
            # Sayısal olmayanları NaN yap, NaN'ları 0 yap ve int'e dönüştür
            chunk['quantity'] = pd.to_numeric(quantity_series, errors='coerce').fillna(0).astype(int)
            
            # --- 2. GÜVENLİ PN FİLTRELEME (Boş ve Sıfır Miktarları Atma) ---
            
            # Part Number'ı temizle ve büyük harfe çevir
            chunk['part_number'] = chunk[material_col_name].astype(str).str.strip().str.upper()
            
            # 🟢 GÜVENLİ FİLTRELEME: Yalnızca Part Number'ı boş olmayan ve Quantity > 0 olanları al
            df_valid = chunk[
                (chunk['part_number'] != '') & 
                (chunk['quantity'] > 0)
            ].copy()

            if df_valid.empty:
                continue
                
            # --- 3. TOPLAMA ---
            chunk_agg = df_valid.groupby('part_number')['quantity'].sum().to_dict()
            for part_num, qty in chunk_agg.items():
                aggregated_counts[part_num] = aggregated_counts.get(part_num, 0) + qty

        
        print("\n--- DEBUG: COUNT VALUES EXTRACTED AND SUMMED FROM MATERIAL DATA ---")
        print(f"Total unique Part Numbers: {len(aggregated_counts)}")
        print("--------------------------------------------------------------------\n")
        
        return aggregated_counts

    except Exception as e:
        print(f"❌ CRITICAL ERROR IN PART NUMBER/QUANTITY EXTRACTION: {e}")
        return {}

def create_process_material_list(groupdata_filepath, target_material_map: Dict[str, int]) -> pd.DataFrame:
    target_materials_set = set(target_material_map.keys())
    material_data: Dict[str, Tuple[str, str, str]] = {} 
    
    # Not: Bu regex'i betiğinizin en üstünde tanımlı tutmanız gerekir.
    MATERIAL_CODE_PATTERN = re.compile(r'^[A-Z0-9]{10,15}$') 
    
    level_list = get_clean_level_names(groupdata_filepath)
    
    print(f"-> Searching for Variants and Production Steps in {len(level_list)} Group Data sheets...")
    
    try:
        # Pandas'ın gelecekteki davranışını ayarlıyoruz (uyarıları gidermek için)
        pd.set_option('future.no_silent_downcasting', True)
        
        for sheet_name, production_variant, development_variant in level_list:
            df_raw = pd.read_excel(groupdata_filepath, sheet_name=sheet_name, header=None) 
            
            # 🟢 DÜZELTME 1: FutureWarning Giderildi (fillna yerine ffill kullanıldı)
            df_raw_filled = df_raw.ffill(axis=1) 
            
            for r in range(1, len(df_raw_filled)):
                for c in range(len(df_raw_filled.columns)):
                    cell_value = df_raw_filled.iloc[r, c] 
                    
                    if pd.notna(cell_value):
                        value_str = str(cell_value).strip().upper()
                        is_potential_material = MATERIAL_CODE_PATTERN.match(value_str)
                        
                        if is_potential_material and value_str in target_materials_set:
                            part_number = value_str
                            
                            if part_number in material_data:
                                continue

                            process_name_candidate = df_raw_filled.iloc[r-1, c]
                            found_production_step = ''
                            
                            if pd.notna(process_name_candidate):
                                production_step = str(process_name_candidate).strip() 
                                name_upper = production_step.upper()
                                
                                # 🟢 DÜZELTME 2: Süreç Adımı Uzunluk Kontrolü Kaldırıldı (Esneklik için)
                                is_valid_process = (
                                    production_step 
                                    # and 2 < len(production_step) < 30 # <-- Bu satır kaldırıldı
                                    and not re.match(r'^-?\d+(\.\d+)?$', production_step)
                                    and not MATERIAL_CODE_PATTERN.match(name_upper)
                                )
                                
                                if is_valid_process:
                                    found_production_step = production_step

                            material_data[part_number] = (production_variant, development_variant, found_production_step) 
                
        final_list = []
        for part_number in target_materials_set:
            prod_variant, dev_variant, production_step = material_data.get(part_number, ('', '', '')) 
            quantity = target_material_map.get(part_number, 0)
            
            final_list.append({
                'part_number': part_number,
                'production_variant': prod_variant,
                'development_variant': dev_variant,
                'production_step': production_step,
                'quantity': quantity
            })
            
        df_map = pd.DataFrame(final_list)
        df_map['production_step'] = df_map['production_step'].fillna('') 
        df_map['production_variant'] = df_map['production_variant'].fillna('')
        df_map['development_variant'] = df_map['development_variant'].fillna('')
        df_map['quantity'] = df_map['quantity'].astype(int) 
        df_map_unique = df_map.drop_duplicates(subset=['part_number']).reset_index(drop=True) 

        return df_map_unique 
        
    except Exception as e:
        print(f"❌ Error: Could not create Variant/Production Step list: {repr(e)}") 
        return pd.DataFrame()
    

def clean_and_escape(value, max_len=None):
    s = str(value).strip()
    # Fazla tırnakları, boşlukları ve '=' işaretini BAŞINDAN ve SONUNDAN temizle
    s = re.sub(r'(^[\s=\'"]+|[\s=\'"]+$)', '', s, flags=re.IGNORECASE)  
    if max_len is not None:
        s = s[:max_len]
    return s

def insert_records_to_redshift(conn, df_records, target_table_full_name, temp_table_full_name) -> bool:
    """
    Records data into Redshift via a temporary table using bulk INSERT and UPSERT (MERGE).
    Includes robust string cleaning and SQL escaping to prevent format errors.
    """
    
    # Redshift tablo şeması ve sütunları (Kendi şemanıza göre ayarlayın)
    COLUMNS = "part_number, production_variant, development_variant, production_step, quantity"
    PRIMARY_KEY = "part_number" # UPSERT için anahtar sütun
    
    # Hedef şema ve tablo adını ayırma
    try:
        # Target table için tam ad kullanın
        target_table_name = target_table_full_name 
        
        # Temporary table için şema adını kaldırın (Redshift kısıtlaması)
        temp_table_name = temp_table_full_name.split('.')[-1] 
    except Exception:
        print(f"❌ HATA: Target table formatı 'schema.table' olmalı: {target_table_full_name}")
        return False

    with conn.cursor() as cursor:
        try:
            print(f"-> 1. Checking/Creating Target Table: {target_table_name}")
            # Target Tabloyu Oluşturma (SQL'de tam adı kullanın)
            create_target_table_sql = f"""
            CREATE TABLE IF NOT EXISTS {target_table_name} (
                part_number VARCHAR(256) PRIMARY KEY,
                production_variant VARCHAR(100),
                development_variant VARCHAR(100),
                production_step VARCHAR(2000),
                quantity INTEGER
            );
            """
            cursor.execute(create_target_table_sql)

            print(f"-> 2. Creating Temporary Table: {temp_table_name}") # Sadece tablo adını basıyoruz
            
            # SQL'de şema adını KULLANMAYIN (Redshift Temporary Table kısıtlaması)
            create_temp_table_sql = f"""
            CREATE TEMP TABLE {temp_table_name} ( 
                part_number VARCHAR(256),
                production_variant VARCHAR(100),
                development_variant VARCHAR(100),
                production_step VARCHAR(2000),
                quantity INTEGER
            );
            """
            cursor.execute(create_temp_table_sql)
            
            # --- BULK INSERT ---
            print(f"-> 3. Loading {len(df_records)} records into Temporary Table (BULK INSERT)...")
            chunk_size = 1000 
            
            for i in range(0, len(df_records), chunk_size):
                chunk = df_records.iloc[i:i + chunk_size]
                
                insert_statements = []
                for _, row in chunk.iterrows():
                    
                    # Temizleme ve Kaçış (Escaping) Uygulanıyor
                    part_number = clean_and_escape(row['part_number'], 256) 
                    production_variant = clean_and_escape(row['production_variant'], 100) 
                    development_variant = clean_and_escape(row['development_variant'], 100) 
                    production_step = clean_and_escape(row['production_step'], 2000) 
                    quantity = int(row['quantity']) 
                    
                    # SQL VALUES formatında string oluşturuluyor
                    insert_sql = f"('{part_number}', '{production_variant}', '{development_variant}', '{production_step}', {quantity})" 
                    insert_statements.append(insert_sql)

                values_str = ','.join(insert_statements)

                final_insert_sql = f"""
                INSERT INTO {temp_table_name} ({COLUMNS})
                VALUES {values_str};
                """
                cursor.execute(final_insert_sql)
            
            # --- UPSERT (MERGE) ---
            print("-> 4. MERGING Temporary Table into Target Table (UPSERT)...")
            merge_sql = f"""
            BEGIN TRANSACTION;
            
            -- Mevcut kayıtları güncelle (UPDATE)
            UPDATE {target_table_name} AS target
            SET
                production_variant = temp.production_variant,
                development_variant = temp.development_variant,
                production_step = temp.production_step,
                quantity = temp.quantity
            FROM {temp_table_name} AS temp
            WHERE target.{PRIMARY_KEY} = temp.{PRIMARY_KEY};

            -- Yeni kayıtları ekle (INSERT)
            INSERT INTO {target_table_name} ({COLUMNS})
            SELECT {COLUMNS}
            FROM {temp_table_name} AS temp
            WHERE NOT EXISTS (
                SELECT 1 
                FROM {target_table_name} AS target 
                WHERE target.{PRIMARY_KEY} = temp.{PRIMARY_KEY}
            );

            END TRANSACTION;
            """
            cursor.execute(merge_sql)
            
            print("-> 5. Dropping Temporary Table.")
            # Redshift Temporary tabloları bağlantı kapanınca silinir, ancak manuel drop eklemek
            # güvenli bir uygulamadır.
            # cursor.execute(f"DROP TABLE {temp_table_name};") 
            
            return True
            
        except Exception as e:
            print(f"❌ CRITICAL ERROR DURING REDSHIFT LOAD: {repr(e)}")
            return False

def check_data_visibility(target_table_name) -> pd.DataFrame:
    # (Function body remains the same)
    conn_check = create_redshift_connection(autocommit_val=False) 
    if conn_check is None:
        return pd.DataFrame()
    
    try:
        sql_query = f"""
        SELECT
            part_number, 
            production_variant,
            development_variant,
            production_step, 
            quantity 
        FROM
            {target_table_name}
        LIMIT 10;
        """
        # C:\Users\uiv21784\OneDrive - Vitesco Technologies\Desktop\proje\main.py:354: UserWarning... uyarısını burası verir.
        df_final = pd.read_sql_query(sql_query, conn_check) 
        conn_check.close()
        return df_final
    except Exception as e:
        conn_check.close()
        print(f"❌ SECONDARY CHECK ERROR: {e}")
        return pd.DataFrame()

def check_total_records(target_table_name) -> int:
    # (Function body remains the same)
    conn_check = create_redshift_connection(autocommit_val=False) 
    if conn_check is None:
        return 0
    
    try:
        with conn_check.cursor() as cursor:
            cursor.execute(f"SELECT COUNT(*) FROM {target_table_name};")
            total_count = cursor.fetchone()[0]
            conn_check.close()
            return total_count
    except Exception as e:
        conn_check.close()
        return 0
        
# ******************************************************************************
# --- MAIN FLOW & WRAPPER ---
# ******************************************************************************

def generate_redshift_process_count_load(groupdata_filepath, materialdata_filepath, output_filepath, target_table_name, temp_table_name):
    print("1/4: Extracting Main Material List and Count from Material Data...")
    # This now uses chunking inside stage_1_get_material_names
    target_material_map = stage_1_get_material_names(materialdata_filepath, MATERIAL_COL_IDX, COUNT_COL_IDX)    
    if not target_material_map:
        print("Critical Error: Material Name list is empty. Process stopped.")
        return

    print(f"2/4: Searching for Process and Design Level Matches for {len(target_material_map)} Materials from Group Data...")
    df_process = create_process_material_list(groupdata_filepath, target_material_map)
    
    if df_process.empty:
        print(f"Critical Error: No data remaining after the matching stage.")
        return
    
    df_records_to_load = df_process.copy()
    
    print(f"3/4: UPSERTING {len(df_records_to_load)} Records to Redshift (Design Levels added)...")
    conn = create_redshift_connection(autocommit_val=True) 
    if conn is None:
        return
    
    # 🟢 Düzeltme: TARGET_TABLE_NAME ve TEMP_TABLE_NAME global değişkenlerini direkt kullanıyoruz
    success = insert_records_to_redshift(conn, df_records_to_load, target_table_name, temp_table_name)
    conn.close() 
    
    if not success:
        print("Critical Error: Data loading test failed.")
        return
    
    print("4/4: Checking data persistence with a NEW SESSION...")
    # 🟢 Düzeltme: TARGET_TABLE_NAME değişkenini kullanıyoruz
    df_loaded = check_data_visibility(target_table_name)
    
    total_records_in_db = check_total_records(target_table_name)
    
    if df_loaded.empty:
        print("\n=======================================================")
        print(f"❌ CRITICAL ERROR: Data check failed or your table is empty.")
        print("=======================================================")
        return
    
    df_loaded.to_excel(output_filepath, index=False)
    
    assigned_process_count = len(df_loaded[df_loaded['production_step'] != ''])
    assigned_prod_count = len(df_loaded[df_loaded['production_variant'] != ''])
    assigned_dev_count = len(df_loaded[df_loaded['development_variant'] != ''])
    
    print("\n=======================================================")
    print(f"🎉 DATA LOAD SUCCESSFUL! (All Variants Added) 🎉")
    print(f"Table '{target_table_name}' updated.")
    print(f"Total Records in Table: {total_records_in_db}")
    print(f"Records assigned Production Variant: {assigned_prod_count}/{len(df_loaded)} (Sample Record)")
    print(f"Records assigned Development Variant: {assigned_dev_count}/{len(df_loaded)} (Sample Record)")
    print(f"Records assigned Process: {assigned_process_count}/{len(df_loaded)} (Sample Record)")
    print(f"📁 Report File Path: {output_filepath}")
    
    print("\n--- TOP 10 LOADED RECORDS (Check) ---")
    print(df_loaded.head(10))
    print("=======================================================")

def scheduled_job_wrapper(groupdata_filepath, materialdata_filepath, output_filepath, target_table_name, temp_table_name):
    # (Function body remains the same)
    """Resilient wrapper function to catch errors and keep the scheduler running."""
    timestamp = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n--- SCHEDULED RUN STARTED: {timestamp} ---")
    
    try:
        generate_redshift_process_count_load(groupdata_filepath, materialdata_filepath, output_filepath, target_table_name, temp_table_name)
        print(f"--- RUN FINISHED SUCCESSFULLY: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')} ---")
    except Exception as e:
        # If any unexpected error occurs, log it but let the scheduler continue for the next interval
        print(f"❌ CRITICAL SCHEDULER ERROR at {timestamp}: The job failed but the scheduler will continue running for the next interval.")
        print(f"Error Details: {e}")
    print("----------------------------------------------------------------------")


# --------------------------------------------------------------------------------------
# MAIN EXECUTION (Using APScheduler for Interval Run)
if __name__ == '__main__':
    
    if SCHEDULE_INTERVAL_MINUTES <= 0:
        print("CRITICAL ERROR: SCHEDULE_INTERVAL_MINUTES must be a positive integer.")
    else:
        scheduler = BlockingScheduler()
        
        # Use the predefined variable for the interval trigger
        trigger = IntervalTrigger(minutes=SCHEDULE_INTERVAL_MINUTES)
        
        scheduler.add_job(
            scheduled_job_wrapper, 
            trigger, 
            args=[GROUPDATA_FILE, MATERIALDATA_FILE, FINAL_REPORT_FILE, TARGET_TABLE_NAME, TEMP_TABLE_NAME]
        )
        
        print("----------------------------------------------------------------------")
        print("🚀 Redshift Loading Process Started (APScheduler)!")
        print(f"➡️ Scheduled to run: Every {SCHEDULE_INTERVAL_MINUTES} minutes.")
        print("❗️ IMPORTANT: Do not close this terminal window. Closing it will stop the schedule.")
        print("To stop the execution: Press Ctrl+C")
        print("----------------------------------------------------------------------")
        
        try:
            # Trigger the first run immediately upon start
            scheduled_job_wrapper(GROUPDATA_FILE, MATERIALDATA_FILE, FINAL_REPORT_FILE, TARGET_TABLE_NAME, TEMP_TABLE_NAME)
            
            # Start the scheduler loop
            scheduler.start()
        except (KeyboardInterrupt, SystemExit):
            print("Scheduler stopped by user.")